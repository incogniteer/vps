#!/usr/bin/env python3

import os
import sys
import time
import datetime
import random
import logging
import smtplib
from string import Template
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from openpyxl import load_workbook
from timeit import default_timer as timer 

def get_contacts(filename, sheets):
  """
  Return 2 lists: names and emails
  read from excel file
  """

  # 客户信息excel路径
  excel_file = filename

  # 加载工作簿
  wb = load_workbook(excel_file)

  # 加载工作表 wb.sheetnames: 返回全部工作表名list
  sheet_names = sheets
  # 客户名存储到names列表，邮箱存储到emails列表
  names = []
  emails = []
  for sheet in sheet_names:
    ws = wb[sheet]

    if ws.max_row > 1:
      # iter_rows 用于递归rows，min_row跳过header
      for row in ws.iter_rows(min_row=2):
        names.append(row[3].value.strip())
        emails.append(row[4].value.strip())

  num_emails = len(emails) # 总收件人数
  return names, emails, num_emails

def read_template(filename):
    """
    Return a Template object comprising the contents of the file specified by filename
    """

    with open(filename, mode='rt', encoding='utf-8') as template:
        template_content = template.read()
    return Template(template_content)

def get_time(s):
  h = s // 3600
  m = s % 3600 // 60
  s = s % 60
  if h > 0:
    return f'{h:.0f}hours,{m:.0f}mins,{s:.3f}s'
  elif m > 0:
    return f'{m:.0f}mins,{s:.3f}s'
  else:
    return f'{s:.3f}s'

def countdown(t, elapsed, num_emails, counter):
    while t:
        min, sec = divmod(t, 60)
        timeformat = f'{min:02d}:{sec:02d}'
        print(f'Next email in {timeformat},{counter} sent in {elapsed}, total: {num_emails}, remaining: {num_emails - counter}', end='\r')
        time.sleep(1)
        t -= 1

# logger函数，用于生成日志，类似tee
def logger():
  logPath = '/home/debian/py/log'
  logFile = 'log'
  logFormatter = '%(asctime)s\t%(levelname)s\t%(message)s'
  date = datetime.datetime.today().strftime('%Y-%m-%d')

  # 默认stderr，重定向到stdout方法：                    
  # import sys + StreamHanlder(sys.stdout) to stdout, default stderr
  logging.basicConfig(level=logging.DEBUG,
                      format=logFormatter,
                      datefmt='%Y-%m-%d %H:%M:%S',
                      handlers=[
                          logging.FileHandler(f'{logPath}/{logFile}-{date}.log', mode='a'),
                          logging.StreamHandler(),
                          ])

  # get the logger
  return logging.getLogger()

logger = logger()

def main(text_msg, html_msg, subject, wb, ws):
    # 登录阿里邮箱，用SMTP_SSL465端口发送邮件
    server = smtplib.SMTP_SSL(host = 'smtp.sinozoc.com', port = 465)
    server.login('sales03@sinozoc.com', 'zc@180730')

    # 加载excel客户信息表，并获取姓名邮箱以及客户数
    workbook =  wb
    sheets = ws
    names, emails, num_emails = get_contacts(workbook, sheets)

    # 程序开始时间
    start = timer() 

    # 邮件发送成功计数器, 必须在for循环外初始化
    counter = 0 

    for name, email in zip(names, emails):
      # create email
      message = MIMEMultipart('alternative')

      # email parameters
      message['From'] = 'Calibur@Sinozoc <sales03@sinozoc.com>'
      message['To'] = email
      message['Cc'] = None
      message['Bcc'] = None
      message['Subject'] = subject

      # MIME text part message
      with open(text_msg, 'r') as fp:
        text_part = Template(fp.read()).substitute(recipient = name.title())

      # MIME html part message
      with open(html_msg, 'r') as fp:
        html_part = Template(fp.read()).substitute(recipient = name.title())

      """
      email.mime.text.MIMEText(_text, _subtype='plain', _charset=None, *, policy=compat32)
      _text string for message payload
      record the MIME types of both parts - text/html and text/html
      turn into MIMEtext objects
      """
      part1 = MIMEText(text_part, 'plain')
      part2 = MIMEText(html_part, 'html')

      """
      attach parts into message container according to RFC 2046, the last part of a multipart message, in this case the HTML message, is best and preferred.
      """
      message.attach(part1)
      message.attach(part2)

      # convert email message to string
      composed = message.as_string()
      server.sendmail('sales03@sinozoc.com', email, composed)
      try:
        server.sendmail('sales03@sinozoc.com', email, composed)
        sys.stdout.write('\033[K') # clear to end of line
        logger.info('Success: ' + name.title() + ':' + email)
        counter += 1

        # finally必须跟try except同indent，下面这段代码放在这里不放在except后
        end = timer() # 程序耗时
        elapsed = get_time(end - start)
        average = get_time((end - start) / counter) 

        # 倒计时跳过最后一项
        if name != names[-1]:
          delay = random.randint(10, 20)
          countdown(delay, elapsed, num_emails, counter)

      # 出错的话就没必要检查运行时间，不计入
      except Exception as e:
        logger.error('Failed: ' + name.title() + ':' + email + ':' + repr(e))

    # 下面不能放在finally会导致每行都输出总信息
    logger.info(f'Total: {num_emails} emails, Sent: {counter} sent in: {elapsed}, average: {average}, Failed: {num_emails - counter} emails.')

    server.quit()

if __name__ == '__main__':
    # 邮件内容
    text_msg ='/home/debian/work/text_msg_es'
    html_msg = '/home/debian/work/html_msg_es'
    # 邮件主题
    subject = 'Saludos - Sinozoc/Calibur'
    # 客户信息表
    wb = "/home/debian/work/customers.xlsx"  
    ws =['test']
    main(text_msg, html_msg, subject, wb, ws)
