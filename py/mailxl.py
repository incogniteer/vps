#!/usr/bin/env py

from openpyxl import load_workbook

def get_contacts(filename):
  """
  Return 2 lists: names and emails
  read from excel file
  """

  # 客户信息excel路径
  excel_file = filename

  # 加载工作簿
  wb = load_workbook(excel_file)

  # 加载工作表 sheetnames 返回全部工作表名list
  ws = wb['Africa']

  names = []
  emails = []
  # iter_rows 用于递归rows，min_row跳过header
  for row in ws.iter_rows(min_row=2):
    names.append(row[3].value.strip())
    emails.append(row[4].value.strip())

  return names, emails

if __name__ == '__main__':
  n, e = get_contacts('/mnt/c/Users/DELL/Desktop/customers.xlsx')
  print(type(n))
  print(n)
  print(*n, sep='\n')
  print(type(e))
  print(*e, sep='\n')
